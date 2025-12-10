# app.py
import os
import io
import json
import base64
import logging
from datetime import datetime
from flask import Flask, render_template, request, jsonify, send_file
from flask_cors import CORS
import google.generativeai as genai
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# -------- Logging --------
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# -------- Flask app --------
app = Flask(__name__, template_folder="templates")
CORS(app)

# -------- Configure API key (AI Studio / API-key only mode) --------
API_KEY = os.getenv("GENAI_API_KEY") or os.getenv("GOOGLE_API_KEY") or ""
if not API_KEY:
    logger.warning("GENAI_API_KEY not set in environment. Set GENAI_API_KEY before running for proper operation.")
else:
    genai.configure(api_key=API_KEY)
    logger.info("Configured genai with API key from environment.")

# In-memory storage - use DB in production
test_cases_storage = {}

# Ordered models to try (API-key-only approach).
# We try them in order and use the first that succeeds.
PREFERRED_MODELS = [
    "models/gemini-2.5-flash",
    "models/gemini-2.5-pro",
    "models/gemini-2.5-flash-image",
    "models/gemini-3-pro-image-preview",
    "models/gemini-flash-latest",
    "models/gemini-2.0-flash",
    # final fallback if none of the above work:
    "models/gemini-2.5-flash"
]

def build_prompt(description: str, expected_count: int = 8):
    return f"""
You are a professional QA engineer. Analyze the UI screenshot (sent separately) and the feature description below.
Generate {expected_count} clear, distinct, and executable UI test cases.

Return ONLY valid JSON: an array of objects. Each object MUST contain exactly:
  tc_id, scenario, steps, expected_result

- tc_id: short ID like TC001
- scenario: one-line test title
- steps: numbered steps as a single string (use '\\n' for newlines)
- expected_result: concise expected outcome

UI Description:
{description}

Return ONLY the JSON array, nothing else.
""".strip()

def attempt_generate_with_models(contents, model_list, override=None):
    """
    Try generate_content with models in model_list in order.
    If override provided, attempt override first.
    Returns (parsed_json_list, used_model, used_fallback_bool, raw_text) or raises Exception.
    """
    tried_models = []
    # if override provided, try it first
    if override:
        try_order = [override] + [m for m in model_list if m != override]
    else:
        try_order = list(model_list)

    last_raw = None
    for idx, model_id in enumerate(try_order):
        tried_models.append(model_id)
        try:
            logger.info("Attempting generation with model: %s", model_id)
            model = genai.GenerativeModel(model_id)
            response = model.generate_content(contents)
            raw_text = getattr(response, "text", None) or str(response)
            raw_text = raw_text.strip()
            last_raw = raw_text

            # strip markdown fences if present
            if raw_text.startswith("```"):
                # remove the first fence line and trailing fence if present
                first_newline = raw_text.find("\n")
                if first_newline != -1:
                    raw_text = raw_text[first_newline+1:]
                if raw_text.endswith("```"):
                    raw_text = raw_text[:-3].strip()

            # attempt to parse JSON
            parsed = json.loads(raw_text)
            if isinstance(parsed, list) and len(parsed) > 0:
                used_fallback = (idx != 0)  # if not the first tried model, mark fallback True
                logger.info("Model %s succeeded and returned %d test cases", model_id, len(parsed))
                return parsed, model_id, used_fallback, raw_text
            else:
                logger.warning("Model %s returned JSON but not a non-empty list. Trying next model.", model_id)
                continue
        except Exception as e:
            # log and try next model
            logger.exception("Model %s failed: %s", model_id, e)
            continue

    # if we fell through all models
    raise RuntimeError(f"All model attempts failed. Last raw output (if any): {last_raw[:1000] if last_raw else 'none'}")

# -------- Routes --------

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate-tests', methods=['POST'])
def generate_tests():
    """
    multipart/form-data expected:
      screenshot (file), description (text), release_version (opt), tester_name (opt),
      run_dates (csv opt), expected_count (opt), model_override (opt)
    """
    try:
        description = request.form.get('description', '').strip()
        release_version = request.form.get('release_version', 'v1.0.0').strip()
        tester_name = request.form.get('tester_name', 'Tester').strip()
        run_dates_raw = request.form.get('run_dates', '').strip()
        expected_count = int(request.form.get('expected_count', '8'))
        model_override = request.form.get('model_override', '').strip() or None

        if not description:
            return jsonify({'success': False, 'error': 'Description is required', 'session_id': None}), 400
        if 'screenshot' not in request.files:
            return jsonify({'success': False, 'error': 'Screenshot file required', 'session_id': None}), 400

        file = request.files['screenshot']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No file selected', 'session_id': None}), 400

        image_bytes = file.read()
        base64_image = base64.b64encode(image_bytes).decode('utf-8')
        fn = file.filename.lower()
        if 'png' in fn:
            mime = 'image/png'
        elif 'gif' in fn:
            mime = 'image/gif'
        elif 'webp' in fn:
            mime = 'image/webp'
        else:
            mime = 'image/jpeg'

        run_dates = [d.strip() for d in run_dates_raw.split(',') if d.strip()] if run_dates_raw else ["17/10/2025","29/10/2025","31/10/2025","21/11/2025"]

        prompt = build_prompt(description, expected_count)
        contents = [prompt, {"mime_type": mime, "data": base64_image}]

        # Try generating with preferred models (uses API-key transport - genai configured above)
        parsed, used_model, used_fallback, raw_text = attempt_generate_with_models(contents, PREFERRED_MODELS, override=model_override)

        # Normalize parsed list into internal structure with runs placeholders
        normalized = []
        for i, tc in enumerate(parsed, start=1):
            if not isinstance(tc, dict):
                tc = {}
            tc_id = tc.get('tc_id') or f"TC{str(i).zfill(3)}"
            scenario = tc.get('scenario') or tc.get('title') or f"Scenario {i}"
            steps = tc.get('steps') or tc.get('test_case_steps') or ''
            expected_result = tc.get('expected_result') or tc.get('expected') or ''
            ts_id = f"TS{str(i).zfill(3)}"
            runs = []
            for rd in run_dates:
                runs.append({'test_date': rd, 'actual_result': '', 'status': 'Not Started', 'bug_id': '', 'commit_id': ''})
            normalized.append({
                'ts_id': ts_id,
                'tc_id': tc_id,
                'scenario': scenario,
                'steps': steps,
                'expected_result': expected_result,
                'release_version': release_version,
                'date_generated': datetime.now().strftime('%Y-%m-%d'),
                'tester_name': tester_name,
                'runs': runs
            })

        session_id = datetime.now().strftime('%Y%m%d%H%M%S')
        test_cases_storage[session_id] = {
            'test_cases': normalized,
            'release_version': release_version,
            'tester_name': tester_name,
            'description': description,
            'run_dates': run_dates,
            'used_model': used_model
        }

        return jsonify({
            'success': True,
            'session_id': session_id,
            'count': len(normalized),
            'test_cases': normalized,
            'used_model': used_model,
            'used_fallback_model': bool(used_fallback)
        }), 200

    except Exception as e:
        logger.exception("generate-tests failed: %s", e)
        return jsonify({'success': False, 'error': str(e), 'session_id': None}), 500

@app.route('/log-results', methods=['POST'])
def log_results():
    """
    Accepts JSON:
    { "session_id":"...", "results":[ { "tc_id":"TC001", "run_index":0, "status":"PASS", "actual_result":"...", "bug_id":"72", "commit_id":"abc" } ] }
    """
    try:
        data = request.get_json(force=True)
        session_id = data.get('session_id')
        results = data.get('results', [])

        if not session_id or session_id not in test_cases_storage:
            return jsonify({'success': False, 'error': 'Invalid session id'}), 400

        session = test_cases_storage[session_id]
        for r in results:
            tc_id = r.get('tc_id')
            run_index = int(r.get('run_index', 0))
            for tc in session['test_cases']:
                if tc['tc_id'] == tc_id:
                    if 0 <= run_index < len(tc['runs']):
                        run = tc['runs'][run_index]
                        run['status'] = r.get('status', run.get('status', 'Not Started'))
                        run['actual_result'] = r.get('actual_result', run.get('actual_result', ''))
                        run['bug_id'] = r.get('bug_id', run.get('bug_id', ''))
                        run['commit_id'] = r.get('commit_id', run.get('commit_id', ''))
                    break

        return jsonify({'success': True}), 200
    except Exception as e:
        logger.exception("log-results failed: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/export-excel/<session_id>', methods=['GET'])
def export_excel(session_id):
    try:
        if session_id not in test_cases_storage:
            return jsonify({'success': False, 'error': 'Invalid session id'}), 400

        session = test_cases_storage[session_id]
        test_cases = session['test_cases']
        run_dates = session.get('run_dates', [])

        wb = Workbook()
        ws = wb.active
        ws.title = "Test Cases"

        # Styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        title_font = Font(bold=True, size=12)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        fill_header = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
        fill_pass = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fill_fail = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        fill_not_started = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        fill_na = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        # Title and metadata
        ws.merge_cells('A1:K1')
        ws['A1'] = f"UI Test Cases - {session['release_version']}"
        ws['A1'].font = title_font
        ws['A2'] = f"Feature / Description: {session['description']}"
        ws['A3'] = f"Generated by: {session['tester_name']} on {test_cases[0]['date_generated'] if test_cases else ''}"
        header_row = 5
        base_headers = ['TS ID', 'Test Scenario', 'TC ID', 'Test Case (Title)', 'Test Case Steps', 'Expected Result']
        per_run_headers = ['Actual Results (AS of Latest Date)', 'Status', 'Test Date', 'Bug ID', 'Commit ID']

        col = 1
        for h in base_headers:
            cell = ws.cell(row=header_row, column=col, value=h); cell.font = header_font; cell.fill = fill_header; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = border; col += 1

        for run_i, rd in enumerate(run_dates):
            for sub in per_run_headers:
                header_text = f"{sub}\n({rd})" if sub == 'Test Date' else f"{sub}\nRun {run_i+1}"
                cell = ws.cell(row=header_row, column=col, value=header_text); cell.font = header_font; cell.fill = fill_header; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True); cell.border = border; col += 1

        ws.freeze_panes = ws['A6']

        row_num = header_row + 1
        for tc in test_cases:
            col = 1
            row_values = [tc.get('ts_id',''), tc.get('scenario',''), tc.get('tc_id',''), tc.get('scenario',''), tc.get('steps',''), tc.get('expected_result','')]
            for val in row_values:
                cell = ws.cell(row=row_num, column=col, value=val); cell.border = border; cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True); col += 1

            for run in tc.get('runs', []):
                cell_actual = ws.cell(row=row_num, column=col, value=run.get('actual_result','')); cell_actual.alignment = Alignment(wrap_text=True); cell_actual.border = border; col += 1
                status_val = (run.get('status') or '').strip().upper()
                cell_status = ws.cell(row=row_num, column=col, value=status_val); cell_status.alignment = Alignment(horizontal='center', vertical='center'); cell_status.border = border
                if status_val in ('PASS','P'): cell_status.fill = fill_pass
                elif status_val in ('FAIL','F'): cell_status.fill = fill_fail
                elif status_val in ('NOT STARTED','NOT_STARTED','TODO',''): cell_status.fill = fill_not_started
                elif status_val in ('NA','N/A'): cell_status.fill = fill_na
                else: cell_status.fill = fill_not_started
                col += 1
                cell_date = ws.cell(row=row_num, column=col, value=run.get('test_date','')); cell_date.border = border; col += 1
                cell_bug = ws.cell(row=row_num, column=col, value=run.get('bug_id','')); cell_bug.border = border; col += 1
                cell_commit = ws.cell(row=row_num, column=col, value=run.get('commit_id','')); cell_commit.border = border; col += 1
            row_num += 1

        total_cols = col
        for i in range(1, total_cols):
            letter = get_column_letter(i)
            if i == 1:
                ws.column_dimensions[letter].width = 12
            elif i in (2,4):
                ws.column_dimensions[letter].width = 30
            elif i == 3:
                ws.column_dimensions[letter].width = 12
            elif i in (5,6):
                ws.column_dimensions[letter].width = 40
            else:
                ws.column_dimensions[letter].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        fname = f"TestCases_{session['release_version']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name=fname)
    except Exception as e:
        logger.exception("export_excel failed: %s", e)
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
